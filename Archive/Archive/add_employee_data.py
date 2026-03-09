import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

# File paths
HR_MAIN_FILE = "HR_Main_DO_NOT_EDIT.xlsx"
TALENT_FILE = "1 Talent Management (2019-2025).xlsx"

# Load data
print("Loading files...")
try:
    # Load Resigned Employees from HR_Main_DO_NOT_EDIT.xlsx
    resigned_df = pd.read_excel(HR_MAIN_FILE, sheet_name="Resigned Employees")
    print(f"✓ Loaded Resigned Employees: {len(resigned_df)} rows")
    print(f"  Columns: {list(resigned_df.columns)}")
    
    # Remove unwanted columns if they exist
    cols_to_drop = [col for col in ['Position/Level_x', 'Age_x', 'Generation_x', 'Tenure.1', 'Age_y', 'Generation_y', 'Position/Level_y'] if col in resigned_df.columns]
    if cols_to_drop:
        resigned_df = resigned_df.drop(columns=cols_to_drop)
        print(f"  Removed duplicate columns: {cols_to_drop}")
    
    # Load Talent Management data
    talent_df = pd.read_excel(TALENT_FILE)
    print(f"✓ Loaded Talent Management: {len(talent_df)} rows")
    print(f"  Columns: {list(talent_df.columns)}")
    
except Exception as e:
    print(f"✗ Error loading files: {e}")
    exit(1)

# Use Employee name as the join key
join_key_resigned = "Employee"
join_key_talent = "Full Name"

print(f"\nMatching employees and adding: Position/Level, Tenure, Age, Generation")

# Get the latest record per employee from talent management
talent_latest = talent_df.sort_values('Fiscal Year').drop_duplicates(
    subset=['Full Name'], 
    keep='last'
)[['Full Name', 'Position/Level', 'Tenured Years', 'Age', 'Generation']]

print(f"  Records from Talent Management: {len(talent_latest)}")

# Create a mapping dictionary
position_map = dict(zip(talent_latest['Full Name'], talent_latest['Position/Level']))
tenure_map = dict(zip(talent_latest['Full Name'], talent_latest['Tenured Years']))
age_map = dict(zip(talent_latest['Full Name'], talent_latest['Age']))
gen_map = dict(zip(talent_latest['Full Name'], talent_latest['Generation']))

# Add columns to resigned_df
resigned_df['Position/Level'] = resigned_df['Employee'].map(position_map)
resigned_df['Tenure'] = resigned_df['Employee'].map(tenure_map)
resigned_df['Age'] = resigned_df['Employee'].map(age_map)
resigned_df['Generation'] = resigned_df['Employee'].map(gen_map)

# Display results
print(f"\n✓ Added columns successfully")
print(f"  Records matched: {resigned_df['Position/Level'].notna().sum()} out of {len(resigned_df)}")
print(f"\nSample data:")
print(resigned_df[['Employee', 'Position/Level', 'Tenure', 'Age', 'Generation']].head(10))

# Save the updated data
print(f"\nSaving updated file...")

# Load the workbook to update the Resigned Employees sheet
wb = openpyxl.load_workbook(HR_MAIN_FILE)
ws = wb["Resigned Employees"]

# Clear existing data completely
ws.delete_rows(1, ws.max_row)

# Write new headers
for col_idx, header in enumerate(resigned_df.columns, 1):
    ws.cell(row=1, column=col_idx, value=header)

# Write new data
for r_idx, row in enumerate(dataframe_to_rows(resigned_df, index=False, header=False), 2):
    for c_idx, value in enumerate(row, 1):
        # Convert NaN to None for Excel
        if pd.isna(value):
            value = None
        ws.cell(row=r_idx, column=c_idx, value=value)

wb.save(HR_MAIN_FILE)
print(f"\n✓ File saved successfully!")
print(f"  Updated: HR_Main_DO_NOT_EDIT.xlsx")
print(f"  Sheet: Resigned Employees")
print(f"  Rows: {len(resigned_df)}")
print(f"  New columns: Position/Level, Tenure, Age, Generation")
