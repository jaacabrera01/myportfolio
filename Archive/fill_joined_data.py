import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

HR_MAIN_FILE = "HR_Main_DO_NOT_EDIT.xlsx"
TALENT_FILE = "1 Talent Management (2019-2025).xlsx"

print("=" * 80)
print("ADDING DATA TO JOINED EMPLOYEES TAB")
print("=" * 80)

# Load joined employees
print(f"\n📋 Loading Joined Employees sheet...")
joined_df = pd.read_excel(HR_MAIN_FILE, sheet_name="Joined Employees")
print(f"✓ Loaded: {len(joined_df)} rows")
print(f"  Columns: {list(joined_df.columns)}")

# Check for null values
print(f"\n🔍 Checking for null/empty values:")
columns_to_check = ['Tenure', 'Age', 'Generation', 'Position/Level']

null_counts = {}
for col in columns_to_check:
    if col in joined_df.columns:
        null_count = joined_df[col].isna().sum()
        null_counts[col] = null_count
        pct = (null_count / len(joined_df)) * 100
        print(f"  {col}: {null_count} empty ({pct:.1f}%)")
    else:
        print(f"  {col}: Column not found - will be added")
        joined_df[col] = None

# Load talent data from all years
print(f"\n📊 Loading Talent Management data (all years)...")
all_talent_data = {}
for year in ['2019', '2020', '2021', '2022', '2023', '2024', '2025']:
    try:
        df = pd.read_excel(TALENT_FILE, sheet_name=year)
        all_talent_data[year] = df
        print(f"✓ {year}: {len(df)} employees")
    except Exception as e:
        print(f"✗ {year}: {e}")

# Combine all talent data and get latest info per employee
print(f"\n🔗 Merging employee data...")
combined_talent = pd.concat(all_talent_data.values(), ignore_index=True)
print(f"  Total records: {len(combined_talent)}")

# Get latest record per employee
latest_talent = combined_talent.sort_values('Fiscal Year').drop_duplicates(
    subset=['Full Name'], 
    keep='last'
)
print(f"  Unique employees: {len(latest_talent)}")

# Create mapping for missing fields
print(f"\n🔄 Filling in missing data...")
matched = 0
not_found = 0

for idx, row in joined_df.iterrows():
    emp_name = row['Employee']
    
    # Find matching employee in talent data
    talent_match = latest_talent[latest_talent['Full Name'] == emp_name]
    
    if len(talent_match) > 0:
        talent_row = talent_match.iloc[0]
        matched += 1
        
        # Fill in tenure
        if pd.isna(joined_df.loc[idx, 'Tenure']) and 'Tenured Years' in talent_row:
            joined_df.loc[idx, 'Tenure'] = talent_row['Tenured Years']
        
        # Fill in age
        if pd.isna(joined_df.loc[idx, 'Age']) and 'Age' in talent_row:
            joined_df.loc[idx, 'Age'] = talent_row['Age']
        
        # Fill in generation
        if pd.isna(joined_df.loc[idx, 'Generation']) and 'Generation' in talent_row:
            joined_df.loc[idx, 'Generation'] = talent_row['Generation']
        
        # Fill in position
        if pd.isna(joined_df.loc[idx, 'Position/Level']) and 'Position/Level' in talent_row:
            joined_df.loc[idx, 'Position/Level'] = talent_row['Position/Level']
    else:
        not_found += 1

print(f"  Matched: {matched} employees")
print(f"  Not found: {not_found} employees")

# Check results
print(f"\n✅ After filling:")
for col in columns_to_check:
    if col in joined_df.columns:
        null_count = joined_df[col].isna().sum()
        pct = (null_count / len(joined_df)) * 100
        print(f"  {col}: {null_count} empty ({pct:.1f}%)")

# Save updated data
print(f"\n💾 Saving updated data...")

wb = openpyxl.load_workbook(HR_MAIN_FILE)
ws = wb["Joined Employees"]

# Clear existing data
ws.delete_rows(1, ws.max_row)

# Write headers
for col_idx, header in enumerate(joined_df.columns, 1):
    ws.cell(row=1, column=col_idx, value=header)

# Write data
for r_idx, row in enumerate(pd.DataFrame(joined_df).values, 2):
    for c_idx, value in enumerate(row, 1):
        if pd.isna(value):
            value = None
        ws.cell(row=r_idx, column=c_idx, value=value)

wb.save(HR_MAIN_FILE)
print(f"✓ File saved!")

# Display sample
print(f"\n📈 Sample data after filling:")
sample_cols = ['Employee', 'Year Joined', 'Tenure', 'Age', 'Generation', 'Position/Level']
print(joined_df[sample_cols].head(15).to_string())

print(f"\n" + "=" * 80)
print("✅ COMPLETE")
print("=" * 80)
print(f"Total employees updated: {matched}")
print(f"Total empty cells filled: {sum(null_counts.values())}")
print("=" * 80)
