import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

HR_MAIN_FILE = "HR_Main_DO_NOT_EDIT.xlsx"
TALENT_FILE = "1 Talent Management (2019-2025).xlsx"

print("=" * 80)
print("CHECKING & FILLING EMPTY CELLS IN RESIGNED EMPLOYEES")
print("=" * 80)

# Load resigned employees
print(f"\n📋 Loading Resigned Employees sheet...")
resigned_df = pd.read_excel(HR_MAIN_FILE, sheet_name="Resigned Employees")
print(f"✓ Loaded: {len(resigned_df)} rows")

# Check for null values
print(f"\n🔍 Checking for null/empty values:")
columns_to_check = ['Tenure', 'Age', 'Generation', 'Position/Level']

null_counts = {}
for col in columns_to_check:
    if col in resigned_df.columns:
        null_count = resigned_df[col].isna().sum()
        null_counts[col] = null_count
        pct = (null_count / len(resigned_df)) * 100
        print(f"  {col}: {null_count} empty ({pct:.1f}%)")
    else:
        print(f"  {col}: Column not found")

# Load talent data from all years
print(f"\n📊 Loading Talent Management data (all years)...")
all_talent_data = {}
for year in ['2019', '2020', '2021', '2022', '2023', '2024', '2025']:
    try:
        df = pd.read_excel(TALENT_FILE, sheet_name=year)
        all_talent_data[year] = df
        print(f"✓ {year}: {len(df)} employees")
    except:
        print(f"✗ {year}: Not found")

# Combine all talent data and get latest info per employee
print(f"\n🔗 Merging employee data...")
combined_talent = pd.concat(all_talent_data.values(), ignore_index=True)
print(f"  Total records: {len(combined_talent)}")

# Get latest record per employee
latest_talent = combined_talent.sort_values('Fiscal Year').drop_duplicates(
    subset=['Full Name'], 
    keep='last'
)
print(f"  Unique employees: {len(latest_talent)}")

# Create mapping for missing fields
print(f"\n🔄 Filling in missing data...")
matched = 0
not_found = 0

for idx, row in resigned_df.iterrows():
    emp_name = row['Employee']
    
    # Find matching employee in talent data
    talent_match = latest_talent[latest_talent['Full Name'] == emp_name]
    
    if len(talent_match) > 0:
        talent_row = talent_match.iloc[0]
        matched += 1
        
        # Fill in missing tenure
        if pd.isna(resigned_df.loc[idx, 'Tenure']) and 'Tenured Years' in talent_row:
            resigned_df.loc[idx, 'Tenure'] = talent_row['Tenured Years']
        
        # Fill in missing age
        if pd.isna(resigned_df.loc[idx, 'Age']) and 'Age' in talent_row:
            resigned_df.loc[idx, 'Age'] = talent_row['Age']
        
        # Fill in missing generation
        if pd.isna(resigned_df.loc[idx, 'Generation']) and 'Generation' in talent_row:
            resigned_df.loc[idx, 'Generation'] = talent_row['Generation']
        
        # Fill in missing position
        if pd.isna(resigned_df.loc[idx, 'Position/Level']) and 'Position/Level' in talent_row:
            resigned_df.loc[idx, 'Position/Level'] = talent_row['Position/Level']
    else:
        not_found += 1

print(f"  Matched: {matched} employees")
print(f"  Not found: {not_found} employees")

# Check results
print(f"\n✅ After filling:")
for col in columns_to_check:
    if col in resigned_df.columns:
        null_count = resigned_df[col].isna().sum()
        pct = (null_count / len(resigned_df)) * 100
        print(f"  {col}: {null_count} empty ({pct:.1f}%)")

# Save updated data
print(f"\n💾 Saving updated data...")

wb = openpyxl.load_workbook(HR_MAIN_FILE)
ws = wb["Resigned Employees"]

# Clear existing data
ws.delete_rows(1, ws.max_row)

# Write headers
for col_idx, header in enumerate(resigned_df.columns, 1):
    ws.cell(row=1, column=col_idx, value=header)

# Write data
for r_idx, row in enumerate(pd.DataFrame(resigned_df).values, 2):
    for c_idx, value in enumerate(row, 1):
        if pd.isna(value):
            value = None
        ws.cell(row=r_idx, column=c_idx, value=value)

wb.save(HR_MAIN_FILE)
print(f"✓ File saved!")

# Display sample
print(f"\n📈 Sample data after filling:")
print(resigned_df[['Employee', 'Year Joined', 'Tenure', 'Age', 'Generation', 'Position/Level']].head(15).to_string())

print(f"\n" + "=" * 80)
print("✅ COMPLETE")
print("=" * 80)
