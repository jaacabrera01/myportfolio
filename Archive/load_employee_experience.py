import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Read source file
source_file = "1 Employee Experience (2019-2025).xlsx"
target_file = "Resigned_Tracker_2019_2025.xlsx"

# Read all sheets from 2019-2025
all_data = []
for year in range(2019, 2026):
    try:
        df_year = pd.read_excel(source_file, sheet_name=str(year))
        df_year['Fiscal Year'] = year
        all_data.append(df_year)
        print(f"Loaded sheet {year}")
    except:
        print(f"Sheet {year} not found, skipping...")

df = pd.concat(all_data, ignore_index=True)

# Count employees by Fiscal Year (assumes each row is an employee with experience)
experience_summary = df.groupby('Fiscal Year').size().reset_index(name='Employee Count')

# Ensure all years 2019-2025 are included
all_years = pd.DataFrame({'Fiscal Year': range(2019, 2026)})
experience_summary = all_years.merge(experience_summary, on='Fiscal Year', how='left').fillna(0)
experience_summary['Employee Count'] = experience_summary['Employee Count'].astype(int)

# Load target workbook and create new sheet
wb = load_workbook(target_file)
if 'Employee Experience' in wb.sheetnames:
    del wb['Employee Experience']

ws = wb.create_sheet('Employee Experience')

# Write data to sheet
for r_idx, row in enumerate(dataframe_to_rows(experience_summary, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        ws.cell(row=r_idx, column=c_idx, value=value)

wb.save(target_file)
print(f"Employee Experience tab created successfully in {target_file}")
print(f"\nSummary:\n{experience_summary}")
