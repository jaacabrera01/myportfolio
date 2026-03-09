import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Read source file
source_file = "1 Talent Management (2019-2025).xlsx"
target_file = "Resigned_Tracker_2019_2025.xlsx"

# Read all sheets from 2019-2025
all_data = []
for year in range(2019, 2026):
    try:
        df_year = pd.read_excel(source_file, sheet_name=str(year))
        df_year['Fiscal Year'] = year
        all_data.append(df_year)
    except:
        print(f"Sheet {year} not found, skipping...")

df = pd.concat(all_data, ignore_index=True)

# Create summary by Fiscal Year and Generation with total count
summary = df.groupby(['Fiscal Year', 'Generation']).size().reset_index(name='Count')

# Pivot to get generations as columns
pivot_summary = summary.pivot(index='Fiscal Year', columns='Generation', values='Count').fillna(0).astype(int)

# Reset index to make Fiscal Year a column
pivot_summary = pivot_summary.reset_index()

# Ensure all years 2019-2025 are included
all_years = pd.DataFrame({'Fiscal Year': range(2019, 2026)})
pivot_summary = all_years.merge(pivot_summary, on='Fiscal Year', how='left').fillna(0)

# Convert generation columns to int
for col in pivot_summary.columns:
    if col != 'Fiscal Year':
        pivot_summary[col] = pivot_summary[col].astype(int)

# Add total column
pivot_summary['Total'] = pivot_summary.drop('Fiscal Year', axis=1).sum(axis=1)

# Load target workbook and create new sheet
wb = load_workbook(target_file)
if 'Talent Management' in wb.sheetnames:
    del wb['Talent Management']
if 'Generation' in wb.sheetnames:
    del wb['Generation']

ws = wb.create_sheet('Generation')

# Write data to sheet
for r_idx, row in enumerate(dataframe_to_rows(pivot_summary, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        ws.cell(row=r_idx, column=c_idx, value=value)

wb.save(target_file)
print(f"Generation tab created successfully in {target_file}")
print(f"\nSummary:\n{pivot_summary}")

# Create Tenure summary
tenure_summary = df.groupby(['Fiscal Year', 'Tenured Years']).size().reset_index(name='Count')

# Pivot to get tenure as columns
pivot_tenure = tenure_summary.pivot(index='Fiscal Year', columns='Tenured Years', values='Count').fillna(0).astype(int)

# Reset index
pivot_tenure = pivot_tenure.reset_index()

# Ensure all years 2019-2025 are included
pivot_tenure = all_years.merge(pivot_tenure, on='Fiscal Year', how='left').fillna(0)

# Convert tenure columns to int
for col in pivot_tenure.columns:
    if col != 'Fiscal Year':
        pivot_tenure[col] = pivot_tenure[col].astype(int)

# Add total column
pivot_tenure['Total'] = pivot_tenure.drop('Fiscal Year', axis=1).sum(axis=1)

# Load workbook again and create Tenure sheet
wb = load_workbook(target_file)
if 'Tenure' in wb.sheetnames:
    del wb['Tenure']

ws_tenure = wb.create_sheet('Tenure')

# Write tenure data to sheet
for r_idx, row in enumerate(dataframe_to_rows(pivot_tenure, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        ws_tenure.cell(row=r_idx, column=c_idx, value=value)

wb.save(target_file)
print(f"Tenure tab created successfully in {target_file}")
print(f"\nTenure Summary:\n{pivot_tenure}")

# Create Age summary
age_summary = df.groupby(['Fiscal Year', 'Age']).size().reset_index(name='Count')

# Pivot to get age as columns
pivot_age = age_summary.pivot(index='Fiscal Year', columns='Age', values='Count').fillna(0).astype(int)

# Reset index
pivot_age = pivot_age.reset_index()

# Ensure all years 2019-2025 are included
pivot_age = all_years.merge(pivot_age, on='Fiscal Year', how='left').fillna(0)

# Convert age columns to int
for col in pivot_age.columns:
    if col != 'Fiscal Year':
        pivot_age[col] = pivot_age[col].astype(int)

# Add total column
pivot_age['Total'] = pivot_age.drop('Fiscal Year', axis=1).sum(axis=1)

# Load workbook again and create Age sheet
wb = load_workbook(target_file)
if 'Age' in wb.sheetnames:
    del wb['Age']

ws_age = wb.create_sheet('Age')

# Write age data to sheet
for r_idx, row in enumerate(dataframe_to_rows(pivot_age, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        ws_age.cell(row=r_idx, column=c_idx, value=value)

wb.save(target_file)
print(f"Age tab created successfully in {target_file}")
print(f"\nAge Summary:\n{pivot_age}")

# Load Employee Experience data
exp_source_file = "1 Employee Experience (2019-2025).xlsx"

# Define the experience categories to count
experience_categories = ['Corporate Culture', 'Job Satisfaction', 'Pay/Benefits', 
                        'Job Content and Design', 'Management', 'Respect', 
                        'Innovation', 'Career', 'Work/Life', 'Leadership', 
                        'Communication', 'Appraisals']

# Read all sheets from Comp_2019 to Comp_2025
exp_data = []
experience_summaries = []

for year in range(2019, 2026):
    try:
        df_exp = pd.read_excel(exp_source_file, sheet_name=f'Comp_{year}', header=None)
        print(f"Loaded Employee Experience sheet Comp_{year}")
        
        year_counts = {'Fiscal Year': year}
        
        # Iterate through rows to find matching categories
        for idx, row in df_exp.iterrows():
            category = str(row[0]).strip() if pd.notna(row[0]) else ""
            
            if category in experience_categories:
                # Get the favorable count from column 1 (second column)
                count_value = row[1] if pd.notna(row[1]) else 0
                year_counts[category] = int(count_value)
        
        experience_summaries.append(year_counts)
        
    except Exception as e:
        print(f"Error loading Comp_{year}: {e}")

# Create DataFrame from extracted summaries
if experience_summaries:
    experience_summary = pd.DataFrame(experience_summaries)
    
    # Add missing categories with 0
    for cat in experience_categories:
        if cat not in experience_summary.columns:
            experience_summary[cat] = 0
    
    # Reorder columns
    cols = ['Fiscal Year'] + experience_categories
    experience_summary = experience_summary[cols]
    
    # Add total column
    experience_summary['Total'] = experience_summary[experience_categories].sum(axis=1)
else:
    # Create empty DataFrame with all years if no data found
    experience_summary = pd.DataFrame({'Fiscal Year': range(2019, 2026)})
    for cat in experience_categories:
        experience_summary[cat] = 0
    experience_summary['Total'] = 0

# Load workbook and create Employee Experience sheet
wb = load_workbook(target_file)
if 'Employee Experience' in wb.sheetnames:
    del wb['Employee Experience']
if 'EE Count' in wb.sheetnames:
    del wb['EE Count']

ws_exp = wb.create_sheet('EE Count')

# Write data to sheet
for r_idx, row in enumerate(dataframe_to_rows(experience_summary, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        ws_exp.cell(row=r_idx, column=c_idx, value=value)

wb.save(target_file)
print(f"EE Count tab created successfully in {target_file}")
print(f"\nEmployee Experience Summary:\n{experience_summary}")

# Create EE Percentage summary
percentage_summaries = []

for year in range(2019, 2026):
    try:
        df_exp = pd.read_excel(exp_source_file, sheet_name=f'Comp_{year}', header=None)
        print(f"Loading percentages from Comp_{year}")
        
        year_percentages = {'Fiscal Year': year}
        
        # Iterate through rows to find matching categories
        for idx, row in df_exp.iterrows():
            category = str(row[0]).strip() if pd.notna(row[0]) else ""
            
            if category in experience_categories:
                # Get the favorable percentage from column 2 (third column)
                percentage_value = row[2] if pd.notna(row[2]) else 0
                year_percentages[category] = float(percentage_value)
        
        percentage_summaries.append(year_percentages)
        
    except Exception as e:
        print(f"Error loading percentages from Comp_{year}: {e}")

# Create DataFrame from extracted percentage summaries
if percentage_summaries:
    percentage_summary = pd.DataFrame(percentage_summaries)
    
    # Add missing categories with 0
    for cat in experience_categories:
        if cat not in percentage_summary.columns:
            percentage_summary[cat] = 0.0
    
    # Reorder columns
    cols = ['Fiscal Year'] + experience_categories
    percentage_summary = percentage_summary[cols]
else:
    # Create empty DataFrame with all years if no data found
    percentage_summary = pd.DataFrame({'Fiscal Year': range(2019, 2026)})
    for cat in experience_categories:
        percentage_summary[cat] = 0.0

# Load workbook and create EE Percentage sheet
wb = load_workbook(target_file)
if 'EE Percentage' in wb.sheetnames:
    del wb['EE Percentage']

ws_pct = wb.create_sheet('EE Percentage')

# Write percentage data to sheet
for r_idx, row in enumerate(dataframe_to_rows(percentage_summary, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        ws_pct.cell(row=r_idx, column=c_idx, value=value)

wb.save(target_file)
print(f"EE Percentage tab created successfully in {target_file}")
print(f"\nEmployee Experience Percentage Summary:\n{percentage_summary}")

# Create Position/Level summary
position_summary = df.groupby(['Fiscal Year', 'Position/Level']).size().reset_index(name='Count')

# Pivot to get positions as columns
pivot_position = position_summary.pivot(index='Fiscal Year', columns='Position/Level', values='Count').fillna(0).astype(int)

# Reset index
pivot_position = pivot_position.reset_index()

# Ensure all years 2019-2025 are included
pivot_position = all_years.merge(pivot_position, on='Fiscal Year', how='left').fillna(0)

# Convert position columns to int
for col in pivot_position.columns:
    if col != 'Fiscal Year':
        pivot_position[col] = pivot_position[col].astype(int)

# Add total column
pivot_position['Total'] = pivot_position.drop('Fiscal Year', axis=1).sum(axis=1)

# Load workbook and create Position/Level sheet
wb = load_workbook(target_file)
if 'Position-Level' in wb.sheetnames:
    del wb['Position-Level']

ws_position = wb.create_sheet('Position-Level')

# Write position data to sheet
for r_idx, row in enumerate(dataframe_to_rows(pivot_position, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        ws_position.cell(row=r_idx, column=c_idx, value=value)

wb.save(target_file)
print(f"Position-Level tab created successfully in {target_file}")
print(f"\nPosition/Level Summary:\n{pivot_position}")
